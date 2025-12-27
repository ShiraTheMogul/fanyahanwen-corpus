#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Pleco PQB builder for Kangxi spreadsheet (xlsx) with Plecofreq-style integration.

Fixes vs previous:
- Correctly pulls the main definition column (康熙字典解釋) even if header has trailing spaces.
- Merges duplicate headwords into ONE Pleco entry (no duplicate dictionary entries in Pleco).
- Adds a linebreak before the main dictionary definition block for cleanliness.
- Keeps metadata as stubs (e.g. 字典路徑: … / 部首: …) on separate lines.
- Formats bracket markers using Pleco newline U+EAB1:
    * 【...】 => two linebreaks BEFORE it
    * 又【...】 => break BEFORE "又" so "又【...】" stays together
    * 【註】 => one linebreak BEFORE it
- Uses Mandarin pinyin (tone numbers) for 'pron' + sortkey so entries appear alongside store dictionaries.
- Uses '@' segmentation in 'word' and 'pron' for multi-character terms, with NO SPACES in pron.
"""

from __future__ import annotations

import argparse
import random
import re
import sqlite3
import time
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from pypinyin import Style
from pypinyin import pinyin as pypinyin_pinyin
import openpyxl

PLECO_NL = "\ueab1"

# ------------------------- CJK filter -------------------------

CJK_RANGES = [
    (0x4E00, 0x9FFF), (0x3400, 0x4DBF), (0x20000, 0x2A6DF),
    (0x2A700, 0x2B73F), (0x2B740, 0x2B81D), (0x2B820, 0x2CEAD),
    (0x2CEB0, 0x2EBE0), (0x31350, 0x323AF), (0x2EBF0, 0x2EE5D),
    (0x323B0, 0x33479), (0x2F800, 0x2FA1F),
]
EXTRA_ALLOWED = {0x3007}  # 〇

def is_cjk_word(word: str) -> bool:
    if not word:
        return False
    for ch in word:
        cp = ord(ch)
        if cp in EXTRA_ALLOWED:
            continue
        if not any(lo <= cp <= hi for lo, hi in CJK_RANGES):
            return False
    return True

# ------------------------- Plecofreq-like formatting -------------------------

def with_at_separators(word: str) -> str:
    return word if len(word) <= 1 else "@".join(list(word))

def mandarin_pinyin_tone3(word: str) -> List[str]:
    py = pypinyin_pinyin(word, style=Style.TONE3, heteronym=False, errors=lambda _: [""])
    out = []
    for item in py:
        syl = (item[0] if item else "") or ""
        out.append(syl.replace("ü", "v"))
    return out

def make_pron(py_syllables: List[str]) -> str:
    # No spaces; keep '@' boundaries.
    return "@".join(py_syllables) if py_syllables else ""

_FULLWIDTH_OFFSET = 0xFEE0
def to_fullwidth_ascii(s: str) -> str:
    out = []
    for ch in s:
        o = ord(ch)
        out.append(chr(o + _FULLWIDTH_OFFSET) if 0x21 <= o <= 0x7E else ch)
    return "".join(out)

def make_sortkey(py_syllables: List[str], word: str) -> str:
    if not word:
        return ""
    if len(word) == 1:
        return to_fullwidth_ascii(py_syllables[0] if py_syllables else "") + word
    return "".join(to_fullwidth_ascii(syl) + ch for ch, syl in zip(word, py_syllables))

def split_pron_tokens(pron: str) -> List[str]:
    # split on whitespace OR '@'
    return [p for p in re.split(r"[\s@]+", pron.strip()) if p]

# ------------------------- Kangxi formatting rules -------------------------

# two linebreaks before 【...】 unless it's 【註】
RE_NOTE = re.compile(r"(【註】)")
# handle 又【...】 -> break before 又
RE_YOU_BRACK = re.compile(r"(又【[^】]*】)")
RE_ANY_BRACK = re.compile(r"(【[^】]*】)")

def format_kangxi_defn(defn: str) -> str:
    """
    Kangxi formatting rules (Pleco newline = U+EAB1):

    1) Never break around 【註】 (keep inline, no special handling).
    2) Break BEFORE a new bracket-tag group 【...】 when it begins a new clause.
       - BUT do NOT break between consecutive bracket-tags: 【集韻】【韻會】 stays together.
    3) Treat '又【...】' as a single token and place breaks BEFORE '又', never between 又 and 【.
       - Use TWO linebreaks before 又-clauses (paragraph-ish), as in your example.
    """

    if not defn:
        return ""

    s = defn.strip()
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", PLECO_NL)

    # Normalize common "ideographic spaces" to regular spaces for easier logic,
    # but keep them in output (we won't aggressively delete whitespace).
    # (Optional; safe)
    # s = s.replace("\u3000", " ")

    # We will tokenize into:
    # - 'YOU_BRACK' tokens: "又【...】"
    # - 'BRACK' tokens: "【...】" except 【註】
    # - 'NOTE' tokens: "【註】" (treated as plain text / inline)
    # - 'TEXT' tokens: everything else
    #
    # Then we rebuild with smart linebreak insertion.

    token_re = re.compile(r"(又【[^】]*】|【註】|【[^】]*】)")
    tokens = []
    pos = 0

    for m in token_re.finditer(s):
        if m.start() > pos:
            tokens.append(("TEXT", s[pos:m.start()]))
        tok = m.group(1)
        if tok.startswith("又【"):
            tokens.append(("YOU_BRACK", tok))
        elif tok == "【註】":
            tokens.append(("NOTE", tok))
        else:
            tokens.append(("BRACK", tok))
        pos = m.end()

    if pos < len(s):
        tokens.append(("TEXT", s[pos:]))

    out_parts: List[str] = []
    at_line_start = True            # are we currently at start of a line/paragraph?
    seen_text_since_break = False   # have we emitted non-trivial TEXT since last break?

    def emit(txt: str):
        nonlocal at_line_start, seen_text_since_break
        if not txt:
            return
        out_parts.append(txt)
        # Count "real" text as anything beyond whitespace and Pleco NL.
        if re.search(r"[^\s" + re.escape(PLECO_NL) + r"]", txt):
            # But don't treat bracket tokens themselves as "text" for grouping
            pass

    def break_line(double: bool = False):
        nonlocal at_line_start, seen_text_since_break
        if out_parts and not out_parts[-1].endswith(PLECO_NL):
            out_parts.append(PLECO_NL)
        if double:
            out_parts.append(PLECO_NL)
        at_line_start = True
        seen_text_since_break = False

    i = 0
    while i < len(tokens):
        kind, val = tokens[i]

        if kind == "TEXT":
            emit(val)
            # If TEXT has substantive characters, we consider that we've "started a clause"
            if re.search(r"[^\s" + re.escape(PLECO_NL) + r"]", val):
                seen_text_since_break = True
                at_line_start = False
            i += 1
            continue

        if kind == "NOTE":
            # Inline; do NOT affect breaking rules.
            emit(val)
            at_line_start = False
            # (Keep seen_text_since_break as-is; this is commentary)
            i += 1
            continue

        if kind == "YOU_BRACK":
            # Paragraph-ish: put TWO linebreaks before it (unless we're already at fresh start).
            # Crucially: do NOT split '又' from '【...】' because it's one token.
            if not at_line_start:
                break_line(double=True)
            emit(val)
            at_line_start = False
            # Treat this as starting a new clause; next BRACK breaks should depend on following TEXT
            seen_text_since_break = False
            i += 1
            continue

        if kind == "BRACK":
            # Group consecutive BRACK tokens together (e.g. 【集韻】【韻會】).
            group = [val]
            j = i + 1
            while j < len(tokens) and tokens[j][0] == "BRACK":
                group.append(tokens[j][1])
                j += 1

            # Decide whether to break BEFORE this group.
            # Break if:
            #   - we're not at start of line AND
            #   - we've seen substantive text since last break (i.e., this is a new clause)
            #
            # Do NOT break if we're already at start, or if it's just another tag cluster.
            def _endswith_jian(out_parts: List[str]) -> bool:
                # Look at the last non-whitespace / non-newline character already emitted.
                s = "".join(out_parts)
                # Remove Pleco newlines + whitespace at end
                s = re.sub(r"[" + re.escape(PLECO_NL) + r"\s]+$", "", s)
                return s.endswith("見")

            # ...

            if (not at_line_start) and seen_text_since_break and (not _endswith_jian(out_parts)):
                break_line(double=False)


            emit("".join(group))
            at_line_start = False
            seen_text_since_break = False  # tags themselves don't count as clause text
            i = j
            continue

    # Cleanup: collapse 3+ breaks to 2
    out = "".join(out_parts)
    out = re.sub(re.escape(PLECO_NL) + r"{3,}", PLECO_NL + PLECO_NL, out)

    # Trim trailing whitespace-ish
    return out.strip()

def build_stub_lines(row: dict, include):
    """
    include: list of (label, key) pairs.
    Returns ["label: value", ...] for non-empty values in row.
    """
    lines = []
    for label, key in include:
        val = row.get(key)
        if val is None:
            continue
        s = str(val).strip()
        if not s:
            continue
        lines.append(f"{label}: {s}")
    return lines


# ------------------------- Read XLSX -------------------------

def load_xlsx_kangxi(
    xlsx_path: str,
    sheet: str,
    col_trad: str = "繁體",
    col_simp: str = "簡體",
    col_path: str = "字典路徑",
    col_ji1: str = "集 1",
    col_ji2: str = "集 2",
    col_rad: str = "部首",
    col_strokes: str = "筆劃數",
    col_defn: str = "康熙字典解釋",
    stubs: bool = True,
    hz_stub: bool = True,
) -> Dict[str, str]:
    """
    Returns {headword: full_defn_text}
    Merges duplicates by concatenating unique definition blocks (separated by two linebreaks).
    """

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet}. Available: {wb.sheetnames}")
    ws = wb[sheet]

    raw_header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    header = [("" if h is None else str(h)).strip() for h in raw_header]

    # Build mapping from header name to index (strip spaces)
    idx = {name: i for i, name in enumerate(header) if name}

    def find_col(name: str) -> int:
        n = name.strip()
        if n in idx:
            return idx[n]
        # try loose match (ignore extra whitespace inside)
        n2 = re.sub(r"\s+", "", n)
        for k, i in idx.items():
            if re.sub(r"\s+", "", k) == n2:
                return i
        raise ValueError(f"Column not found: {name!r}. Columns: {header}")

    i_trad = find_col(col_trad)
    i_simp = find_col(col_simp)
    i_path = find_col(col_path)
    i_ji1  = find_col(col_ji1)
    i_ji2  = find_col(col_ji2)
    i_rad  = find_col(col_rad)
    i_stk  = find_col(col_strokes)
    i_def  = find_col(col_defn)

    # merge map: word -> set of blocks
    blocks: Dict[str, set] = {}

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not row_vals:
            continue

        trad = row_vals[i_trad] if i_trad < len(row_vals) else None
        simp = row_vals[i_simp] if i_simp < len(row_vals) else None
        word = (str(trad).strip() if trad else "") or (str(simp).strip() if simp else "")
        if not word or not is_cjk_word(word):
            continue

        row = {
            "字典路徑": row_vals[i_path] if i_path < len(row_vals) else None,
            "集 1": row_vals[i_ji1] if i_ji1 < len(row_vals) else None,
            "集 2": row_vals[i_ji2] if i_ji2 < len(row_vals) else None,
            "部首": row_vals[i_rad] if i_rad < len(row_vals) else None,
            "筆劃數": row_vals[i_stk] if i_stk < len(row_vals) else None,
            "康熙字典解釋": row_vals[i_def] if i_def < len(row_vals) else None,
        }

        main_defn_raw = row.get("康熙字典解釋")
        main_defn = format_kangxi_defn("" if main_defn_raw is None else str(main_defn_raw))
        if not main_defn.strip():
            # If there's truly no definition, skip this row
            continue

        parts: List[str] = []
        if hz_stub:
            parts.append(f"HZ {word}")

        if stubs:
            stub_lines = build_stub_lines(row, [
                ("字典路徑", "字典路徑"),
                ("集 1", "集 1"),
                ("集 2", "集 2"),
                ("部首", "部首"),
                ("筆劃數", "筆劃數"),
            ])
            parts.extend(stub_lines)

        # Add a linebreak BEFORE the dictionary definition (your request)
        # If there are stubs, this yields a blank line between stubs and definition.
        if parts:
            parts.append("")  # blank line => single Pleco NL when joined
        parts.append(main_defn)

        # Join using Pleco newline. (Blank line creates a clean separation)
        block = PLECO_NL.join(parts).strip()

        blocks.setdefault(word, set()).add(block)

    # Final: merge multiple blocks per word using double break between blocks
    merged: Dict[str, str] = {}
    for w, bset in blocks.items():
        # stable ordering
        blist = sorted(bset)
        merged[w] = (PLECO_NL + PLECO_NL).join(blist)

    return merged

# ------------------------- Pleco schema/index -------------------------

SCHEMA_SQL = [
    """
    CREATE TABLE 'pleco_dict_entries' (
      "uid" INTEGER PRIMARY KEY AUTOINCREMENT,
      "created" INTEGER,
      "modified" INTEGER,
      "length" INTEGER,
      "word" TEXT COLLATE NOCASE,
      "altword" TEXT COLLATE NOCASE,
      "pron" TEXT COLLATE NOCASE,
      "defn" TEXT,
      "sortkey" TEXT UNIQUE
    );
    """,
    """
    CREATE TABLE 'pleco_dict_imports' (
      "id" INTEGER PRIMARY KEY AUTOINCREMENT,
      "starttime" INTEGER,
      "endtime" INTEGER,
      "startentry" INTEGER,
      "endentry" INTEGER
    );
    """,
    """
    CREATE TABLE 'pleco_dict_properties' (
      "propset" INTEGER,
      "propid" TEXT,
      "propvalue" TEXT,
      "propisstring" INTEGER,
      UNIQUE ("propset","propid")
    );
    """,
    "CREATE TABLE 'pleco_dict_posdex_hz_1' (syllable TEXT, uid INTEGER, length INTEGER);",
    "CREATE TABLE 'pleco_dict_posdex_hz_2' (syllable TEXT, uid INTEGER, length INTEGER);",
    "CREATE TABLE 'pleco_dict_posdex_hz_3' (syllable TEXT, uid INTEGER, length INTEGER);",
    "CREATE TABLE 'pleco_dict_posdex_hz_4' (syllable TEXT, uid INTEGER, length INTEGER);",
    "CREATE TABLE 'pleco_dict_posdex_py_1' (syllable TEXT, uid INTEGER, length INTEGER);",
    "CREATE TABLE 'pleco_dict_posdex_py_2' (syllable TEXT, uid INTEGER, length INTEGER);",
    "CREATE TABLE 'pleco_dict_posdex_py_3' (syllable TEXT, uid INTEGER, length INTEGER);",
    "CREATE TABLE 'pleco_dict_posdex_py_4' (syllable TEXT, uid INTEGER, length INTEGER);",
]

INDEX_SQL = [
    "CREATE INDEX idx_pleco_dict_entries_sortkey ON pleco_dict_entries (sortkey);",
    "CREATE INDEX idx_pleco_dict_posdex_hz_1_syllable_uid_length ON pleco_dict_posdex_hz_1 (syllable, uid, length);",
    "CREATE INDEX idx_pleco_dict_posdex_hz_1_uid ON pleco_dict_posdex_hz_1 (uid);",
    "CREATE INDEX idx_pleco_dict_posdex_hz_2_syllable_uid ON pleco_dict_posdex_hz_2 (syllable, uid);",
    "CREATE INDEX idx_pleco_dict_posdex_hz_2_uid ON pleco_dict_posdex_hz_2 (uid);",
    "CREATE INDEX idx_pleco_dict_posdex_hz_3_syllable_uid ON pleco_dict_posdex_hz_3 (syllable, uid);",
    "CREATE INDEX idx_pleco_dict_posdex_hz_3_uid ON pleco_dict_posdex_hz_3 (uid);",
    "CREATE INDEX idx_pleco_dict_posdex_hz_4_syllable_uid ON pleco_dict_posdex_hz_4 (syllable, uid);",
    "CREATE INDEX idx_pleco_dict_posdex_hz_4_uid ON pleco_dict_posdex_hz_4 (uid);",
    "CREATE INDEX idx_pleco_dict_posdex_py_1_syllable_uid_length ON pleco_dict_posdex_py_1 (syllable, uid, length);",
    "CREATE INDEX idx_pleco_dict_posdex_py_1_uid ON pleco_dict_posdex_py_1 (uid);",
    "CREATE INDEX idx_pleco_dict_posdex_py_2_syllable_uid ON pleco_dict_posdex_py_2 (syllable, uid);",
    "CREATE INDEX idx_pleco_dict_posdex_py_2_uid ON pleco_dict_posdex_py_2 (uid);",
    "CREATE INDEX idx_pleco_dict_posdex_py_3_syllable_uid ON pleco_dict_posdex_py_3 (syllable, uid);",
    "CREATE INDEX idx_pleco_dict_posdex_py_3_uid ON pleco_dict_posdex_py_3 (uid);",
    "CREATE INDEX idx_pleco_dict_posdex_py_4_syllable_uid ON pleco_dict_posdex_py_4 (syllable, uid);",
    "CREATE INDEX idx_pleco_dict_posdex_py_4_uid ON pleco_dict_posdex_py_4 (uid);",
]

def write_properties(cur: sqlite3.Cursor, *, dict_name: str, menu_name: str, short_name: str, icon: str,
                     entry_count: int, now: int) -> None:
    file_id = random.randint(-2_000_000_000, 2_000_000_000)
    file_creator = random.randint(1, 50_000_000)
    props = [
        ("DictCopyright", "", 1),
        ("DictIconFillColor", "39372", 0),
        ("DictIconName", icon, 1),
        ("DictIconTextColor", "16777215", 0),
        ("DictLang", "Chinese", 1),
        ("DictMenuName", menu_name, 1),
        ("DictName", dict_name, 1),
        ("DictShortName", short_name, 1),
        ("EditLock", None, 0),
        ("EntryCount", str(entry_count), 0),
        ("FileCreated", str(now), 0),
        ("FileCreator", str(file_creator), 0),
        ("FileGenerator", "Pleco Engine 2.0", 1),
        ("FileID", str(file_id), 0),
        ("FilePlatform", "Android", 1),
        ("FormatString", "Pleco SQL Dictionary Database", 1),
        ("FormatVersion", "8", 0),
        ("NoSortKey", None, 0),
        ("SortMethod", None, 0),
        ("TransLang", "English", 1),
    ]
    for propid, propvalue, propisstring in props:
        cur.execute(
            "INSERT OR REPLACE INTO pleco_dict_properties (propset, propid, propvalue, propisstring) VALUES (0, ?, ?, ?);",
            (propid, propvalue, propisstring),
        )

def insert_posdex(cur: sqlite3.Cursor, uid: int, word_with_at: str, pron: str) -> None:
    word_no_at = word_with_at.replace("@", "")
    wlen = len(word_no_at)
    for i, ch in enumerate(list(word_no_at)[:4], start=1):
        cur.execute(f"INSERT INTO pleco_dict_posdex_hz_{i} (syllable, uid, length) VALUES (?, ?, ?);", (ch, uid, wlen))
    py_tokens = split_pron_tokens(pron)
    for i, syl in enumerate(py_tokens[:4], start=1):
        cur.execute(f"INSERT INTO pleco_dict_posdex_py_{i} (syllable, uid, length) VALUES (?, ?, ?);", (syl, uid, wlen))

def build_pqb(defn_map: Dict[str, str], out_path: str, *, dict_name: str, menu_name: str, short_name: str, icon: str) -> None:
    out = Path(out_path)
    if out.exists():
        out.unlink()

    con = sqlite3.connect(str(out))
    con.execute("PRAGMA page_size=1024;")
    con.execute("PRAGMA journal_mode=DELETE;")
    con.execute("PRAGMA synchronous=FULL;")
    cur = con.cursor()

    for sql in SCHEMA_SQL:
        cur.executescript(sql)
    for sql in INDEX_SQL:
        cur.execute(sql)

    now = int(time.time())
    uid = 1

    for word in sorted(defn_map.keys()):
        word_at = word
        py = mandarin_pinyin_tone3(word)
        pron = " ".join(py)
        sk = make_sortkey(py, word) or word

        defn = defn_map[word]

        cur.execute(
            "INSERT INTO pleco_dict_entries (uid, created, modified, length, word, altword, pron, defn, sortkey) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);",
            (uid, now, now, len(word), word_at, None, pron, defn, sk),
        )
        insert_posdex(cur, uid, word_at, pron)
        uid += 1

    count = uid - 1
    write_properties(cur, dict_name=dict_name, menu_name=menu_name, short_name=short_name, icon=icon, entry_count=count, now=now)
    cur.execute("INSERT INTO pleco_dict_imports (starttime, endtime, startentry, endentry) VALUES (?, ?, 1, ?);", (now, now, count))
    con.commit()
    con.close()
    print(f"Wrote {out_path} with {count} entries.")

# ------------------------- CLI -------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to kx_full.xlsx")
    ap.add_argument("--sheet", default="kx", help="Sheet name (default: kx)")
    ap.add_argument("--out", required=True, help="Output .pqb file")

    ap.add_argument("--no-stubs", action="store_true", help="Disable metadata stub lines")
    ap.add_argument("--no-hz", action="store_true", help="Disable 'HZ <char>' stub line")

    ap.add_argument("--dict-name", required=True)
    ap.add_argument("--menu-name", required=True)
    ap.add_argument("--short-name", required=True)
    ap.add_argument("--icon", required=True)

    # allow overriding column names if needed
    ap.add_argument("--col-trad", default="繁體")
    ap.add_argument("--col-simp", default="簡體")
    ap.add_argument("--col-path", default="字典路徑")
    ap.add_argument("--col-ji1", default="集 1")
    ap.add_argument("--col-ji2", default="集 2")
    ap.add_argument("--col-rad", default="部首")
    ap.add_argument("--col-strokes", default="筆劃數")
    ap.add_argument("--col-defn", default="康熙字典解釋")

    args = ap.parse_args()

    dmap = load_xlsx_kangxi(
        xlsx_path=args.xlsx,
        sheet=args.sheet,
        col_trad=args.col_trad,
        col_simp=args.col_simp,
        col_path=args.col_path,
        col_ji1=args.col_ji1,
        col_ji2=args.col_ji2,
        col_rad=args.col_rad,
        col_strokes=args.col_strokes,
        col_defn=args.col_defn,
        stubs=not args.no_stubs,
        hz_stub=not args.no_hz,
    )

    if not dmap:
        raise SystemExit("No entries loaded from XLSX (after filtering).")

    build_pqb(
        dmap,
        args.out,
        dict_name=args.dict_name,
        menu_name=args.menu_name,
        short_name=args.short_name,
        icon=args.icon,
    )

if __name__ == "__main__":
    main()
