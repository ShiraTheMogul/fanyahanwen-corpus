#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
shuowen_folder_to_xlsx.py

Reads every *.txt in the current folder (or a specified folder),
extracts Shuowen-style categories and entries, and writes an .xlsx.

Key parsing rules (easy to edit for other works):
- Category line: ends with '部' (e.g. '目部', '角部')
- Entry line: contains a colon '：' (fullwidth) or ':' (ascii)
  and the "character/headword" is the text before the first colon.

Output columns:
- source_file
- row_in_source (1-based)
- shuowen_category
- character
- entry
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def iter_text_files(folder: Path) -> Iterable[Path]:
    # Pattern you can reuse elsewhere:
    # For every file in Y that matches X...
    # Here: X = "*.txt", Y = folder
    yield from sorted(folder.glob("*.txt"))


def is_category_line(line: str) -> bool:
    # Shuowen categories look like "目部", "角部", etc.
    # Be conservative: short-ish and ends with 部.
    s = line.strip()
    return s.endswith("部") and len(s) <= 10


def split_entry_line(line: str) -> Optional[Tuple[str, str]]:
    """
    Returns (headword, entry_text) or None if not an entry line.

    Accepts both fullwidth '：' and ASCII ':'.
    Uses the FIRST colon only.
    """
    s = line.strip()
    if not s:
        return None

    # Prefer fullwidth colon first (common in your files).
    if "：" in s:
        head, rest = s.split("：", 1)
    elif ":" in s:
        head, rest = s.split(":", 1)
    else:
        return None

    head = head.strip()
    rest = rest.strip()

    # Shuowen entry headwords are usually 1 char, but sometimes may be more.
    # Keep it permissive; just require a non-empty head and non-empty body.
    if not head or not rest:
        return None

    return head, rest


def folder_to_rows(folder: Path) -> list[dict]:
    rows: list[dict] = []

    for txt_path in iter_text_files(folder):
        current_category = None

        # Read as UTF-8; if your pipeline ever outputs BOM or weird encodings,
        # this is the one place you'd change it.
        lines = txt_path.read_text(encoding="utf-8").splitlines()

        for idx, raw in enumerate(lines, start=1):
            line = raw.strip()
            if not line:
                continue

            # Skip metadata/comment lines
            if line.startswith("#"):
                continue

            # 1) Category header
            if is_category_line(line):
                current_category = line  # store literally "目部"
                continue

            # 2) Entry line
            split = split_entry_line(line)
            if split is None:
                continue

            character, entry = split

            rows.append(
                {
                    "source_file": txt_path.name,
                    "row_in_source": idx,
                    "shuowen_category": current_category,
                    "character": character,
                    "entry": entry,
                }
            )

    return rows


def autosize_columns(ws) -> None:
    # Simple width heuristic so the sheet isn't painful to read.
    # Pattern: For each column, compute max string length, set width.
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)


def write_xlsx(rows: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "shuowen"

    headers = ["source_file", "row_in_source", "shuowen_category", "character", "entry"]
    ws.append(headers)

    for r in rows:
        ws.append([r.get(h) for h in headers])

    autosize_columns(ws)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert a folder of Shuowen TXT files into an XLSX table."
    )
    parser.add_argument(
        "--folder",
        default=".",
        help="Folder containing .txt files (default: current directory).",
    )
    parser.add_argument(
        "--out",
        default="shuowen.xlsx",
        help="Output xlsx filename (default: shuowen.xlsx).",
    )
    args = parser.parse_args()

    folder = Path(args.folder).resolve()
    out_path = Path(args.out).resolve()

    rows = folder_to_rows(folder)
    write_xlsx(rows, out_path)

    print(f"[ok] folder={folder}")
    print(f"[ok] files_found={len(list(iter_text_files(folder)))}")
    print(f"[ok] rows_written={len(rows)}")
    print(f"[ok] output={out_path}")


if __name__ == "__main__":
    main()
